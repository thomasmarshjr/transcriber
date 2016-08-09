from openpyxl import Workbook, worksheet
from openpyxl import load_workbook
import urllib2, html2text
import re

def getTxtFiles(txt):
    if conformed_submission_type == "10-K":
        BD_sec_loc1 = txt.find("\nBusiness", 10000)
        BD_sec_loc2 = txt.find("\nRisk Factors", BD_sec_loc1)
        BD_sec = txt[BD_sec_loc1:BD_sec_loc2]
        BD_txt = open(txt_file_name + " BD", 'w+')
        BD_txt.write(BD_sec)
        BD_txt.close()

        RF_sec_loc1 = txt.find("\nRisk Factors", 10000)
        RF_sec_loc2 = txt.find("\nUnresolved Staff Comments", RF_sec_loc1)
        RF_sec = txt[RF_sec_loc1:RF_sec_loc2]
        RF_txt = open(txt_file_name + " RF.txt", 'w+')
        RF_txt.write(RF_sec)
        RF_txt.close()

        LP_sec_loc1 = txt.find("\nLegal Proceedings", 10000)
        LP_sec_loc2 = txt.find("\nMine Safety Disclosures", LP_sec_loc1)
        LP_sec = txt[LP_sec_loc1:LP_sec_loc2]
        LP_txt = open(txt_file_name + " LP.txt", 'w+')
        LP_txt.write(LP_sec)
        LP_txt.close()

        MDA_sec_loc1 = txt.find("\nManagement", 10000)
        MDA_sec_loc2 = txt.find("\nQuantitative and Qualitative", MDA_sec_loc1)
        MDA_sec = txt[MDA_sec_loc1:MDA_sec_loc2]
        MDA_txt = open(txt_file_name + " MDA.txt", 'w+')
        MDA_txt.write(MDA_sec)
        MDA_txt.close()

        NTF_sec_loc1 = txt.find("\nFinancial Statements", 10000)
        NTF_sec_loc2 = txt.find("\nChanges in and Disagreements", NTF_sec_loc1)
        NTF_sec = txt[NTF_sec_loc1:NTF_sec_loc2]
        NTF_txt = open(txt_file_name + " NTF.txt", 'w+')
        NTF_txt.write(NTF_sec)
        NTF_txt.close()
    elif conformed_submission_type == "10-Q":
        NTF_sec_loc1 = txt.find("\nFinancial Statements", 10000)
        if NTF_sec_loc1 == -1:
            NTF_sec_loc1 = txt.find("\nConsolidated Financial Statements", 10000)
        NTF_sec_loc2 = txt.find("\nManagement", NTF_sec_loc1)
        NTF_sec = txt[NTF_sec_loc1:NTF_sec_loc2]
        NTF_txt = open(txt_file_name + " NTF.txt", 'w+')
        NTF_txt.write(NTF_sec)
        NTF_txt.close()

        MDA_sec_loc1 = txt.find("\nManagement's Discussion and Analysis", 10000)
        if MDA_sec_loc1 == -1:
            MDA_sec_loc1 = txt.find("\nManagements Discussion and Analysis")
        MDA_sec_loc2 = txt.find("\nQuantitative and Qualitative", 10000)
        MDA_sec = txt[MDA_sec_loc1:MDA_sec_loc2]
        MDA_txt = open(txt_file_name + " MDA.txt", 'w+')
        MDA_txt.write(MDA_sec)
        MDA_txt.close()

        LP_sec_loc1 = txt.find("\nLegal Proceedings", 10000)
        LP_sec_loc2 = txt.find("\nRisk Factors", LP_sec_loc1)
        LP_sec = txt[LP_sec_loc1:LP_sec_loc2]
        LP_txt = open(txt_file_name + " LP.txt", 'w+')
        LP_txt.write(LP_sec)
        LP_txt.close()

        RF_sec_loc1 = txt.find("\nRisk Factors", 10000)
        RF_sec_loc2 = txt.find("\nUnregistered Sales of Equity", RF_sec_loc1)
        if RF_sec_loc2 == -1:
            RF_sec_loc2 = txt.find("\nExhibits", RF_sec_loc1)
        RF_sec = txt[RF_sec_loc1:RF_sec_loc2]
        RF_txt = open(txt_file_name + " RF.txt", 'w+')
        RF_txt.write(RF_sec)
        RF_txt.close()

def remove_non_ascii_1(text):
    return ''.join(i for i in text if ord(i) < 128)

wb = Workbook()
ws2 = wb.active
ws2.title = "Financial Info"
ws3 = wb.create_sheet("SEC Filing Metadata")


wb1 = load_workbook('SEC_Metadata.xlsx', use_iterators = True)
ws1 = wb1.get_sheet_by_name("SEC_Metadata")

i=2
x=2

ws3.cell(column=1, row=1, value='Company_Name')
ws3.cell(column=2, row=1, value='Submission_Type')
ws3.cell(column=3, row=1, value='Filing_Date')
ws3.cell(column=4, row=1, value='Period_of_Report')
ws3.cell(column=5, row=1, value='Fiscal_Year_End')
ws3.cell(column=6, row=1, value='Accession_Number')
ws3.cell(column=7, row=1, value='CIK')
ws3.cell(column=8, row=1, value='SIC')
ws3.cell(column=9, row=1, value='IRS_Number')
ws3.cell(column=10, row=1, value='State_Incorporated')
ws3.cell(column=11, row=1, value='Street')
ws3.cell(column=12, row=1, value='City')
ws3.cell(column=13, row=1, value='State')
ws3.cell(column=14, row=1, value='Zipcode')
ws3.cell(column=15, row=1, value='Text_URL')
ws3.cell(column=16, row=1, value='HTML_Link')
ws3.cell(column=17, row=1, value='MDA_Link')
ws3.cell(column=18, row=1, value='NTF_Link')
ws3.cell(column=19, row=1, value='LP_Link')
ws3.cell(column=20, row=1, value='RF_Link')
ws3.cell(column=21, row=1, value='BD_Link')

ws2.cell(column=1, row=1, value='adsh')
ws2.cell(column=2, row=1, value='metric')
ws2.cell(column=3, row=1, value='tag')
ws2.cell(column=4, row=1, value='version')
ws2.cell(column=5, row=1, value='coreg')
ws2.cell(column=6, row=1, value='ddate')
ws2.cell(column=7, row=1, value='qtrs')
ws2.cell(column=8, row=1, value='uom')
ws2.cell(column=9, row=1, value='value')
ws2.cell(column=10, row=1, value='footnote')

for row in ws1.rows:
    for cell in row:
        url_link = cell.value

        response = urllib2.urlopen(url_link)
        html = response.read()
        sec_url = 'https://www.sec.gov/Archives/edgar/data/'

        accession_number_loc1 = html.find("ACCESSION NUMBER:") + 19
        accession_number_loc2 = html.find("\n", accession_number_loc1)
        accession_number = html[accession_number_loc1:accession_number_loc2]

        print accession_number

        conformed_submission_type_loc1 = html.find("CONFORMED SUBMISSION TYPE:") + 27
        conformed_submission_type_loc2 = html.find("\n", conformed_submission_type_loc1)
        conformed_submission_type = html[conformed_submission_type_loc1:conformed_submission_type_loc2]

        conformed_period_of_report_loc1 = html.find("CONFORMED PERIOD OF REPORT:") + 28
        conformed_period_of_report_loc2 = html.find("\n", conformed_period_of_report_loc1)
        conformed_period_of_report = html[conformed_period_of_report_loc1:conformed_period_of_report_loc2]

        filing_date_loc1 = html.find("FILED AS OF DATE:") + 19
        filing_date_loc2 = html.find("\n", filing_date_loc1)
        filing_date = html[filing_date_loc1:filing_date_loc2]

        company_name_loc1 = html.find("COMPANY CONFORMED NAME:") + 26
        company_name_loc2 = html.find("\n", company_name_loc1)
        company_name = html[company_name_loc1:company_name_loc2]

        cik_number_loc1 = html.find("CENTRAL INDEX KEY:") + 21
        cik_number_loc2 = html.find("\n", cik_number_loc1)
        cik_number = html[cik_number_loc1:cik_number_loc2]

        cik_url = url_link[url_link.find("data/") + 5:url_link.find("/", url_link.find("data/") + 5)]
        text_url = sec_url + cik_url + "/" + accession_number + ".txt"

        file_name_loc1 = html.find("<FILENAME>") + 10
        file_name_loc2 = html.find("\n", file_name_loc1)
        file_name = html[file_name_loc1:file_name_loc2]

        sic_group_loc1 = html.find("STANDARD INDUSTRIAL CLASSIFICATION:") + 36
        sic_group_loc2 = html.find("\n", sic_group_loc1)
        sic_group = html[sic_group_loc1:sic_group_loc2]

        irs_number_loc1 = html.find("IRS NUMBER:") + 15
        irs_number_loc2 = html.find("\n", irs_number_loc1)
        irs_number = html[irs_number_loc1:irs_number_loc2]

        state_incorporated_loc1 = html.find("STATE OF INCORPORATION:") + 26
        state_incorporated_loc2 = html.find("\n", state_incorporated_loc1)
        state_incorporated = html[state_incorporated_loc1:state_incorporated_loc2]

        fiscal_year_end_loc1 = html.find("FISCAL YEAR END:") + 19
        fiscal_year_end_loc2 = html.find("\n", fiscal_year_end_loc1)
        fiscal_year_end = html[fiscal_year_end_loc1:fiscal_year_end_loc2]

        street_loc1 = html.find("STREET 1:") + 11
        street_loc2 = html.find("\n", street_loc1)
        street = html[street_loc1:street_loc2]

        city_loc1 = html.find("CITY:") + 8
        city_loc2 = html.find("\n", city_loc1)
        city = html[city_loc1:city_loc2]

        state_loc1 = html.find("STATE:") + 9
        state_loc2 = html.find("\n", state_loc1)
        state = html[state_loc1:state_loc2]

        zip_loc1 = html.find("ZIP:") + 7
        zip_loc2 = html.find("\n", zip_loc1)
        zipcode = html[zip_loc1:zip_loc2]

        html_link = sec_url + cik_url + "/" + accession_number.replace("-", "") + "/" + file_name

        response = urllib2.urlopen(html_link)
        html = response.read()
        fixed = html2text.html2text(html)
        txt_file_name = company_name.replace('/', '') + ' ' + conformed_submission_type + ' ' + filing_date
        txt_file = open(txt_file_name+'.txt', 'w+')
        txt_file.write (remove_non_ascii_1(fixed))
        txt_file.close()

        #sec_txt = open(txt_file_name+'.txt', 'r')
        #txt = sec_txt.read()
        #getTxtFiles(txt)



        MDA_url_link = ""
        NTF_url_link = ""
        RF_url_link = ""
        LP_url_link = ""
        BD_url_link = ""

        if conformed_submission_type == "10-Q":
            MDA_href_loc1 = html.find("Management&")
            MDA_href_loc2 = html.find("href=", MDA_href_loc1 - 90)
            if MDA_href_loc2 > MDA_href_loc1:
                MDA_href_loc2 = -1
            if MDA_href_loc2 == -1:
                MDA_href_loc2 = html.find("HREF=", MDA_href_loc1 - 90)
            MDA_href_loc3 = html.find('"', MDA_href_loc2 + 8)
            MDA_href = html[MDA_href_loc2 + 6: MDA_href_loc3]
            MDA_url_link = html_link + MDA_href

            NTF_href_loc1 = html.find("Notes to Con")
            NTF_href_loc2 = html.find("href=", NTF_href_loc1 - 90)
            if NTF_href_loc2 > NTF_href_loc1:
                NTF_href_loc2 = -1
            if NTF_href_loc2 == -1:
                NTF_href_loc2 = html.find("HREF=", NTF_href_loc1 - 90)
            NTF_href_loc3 = html.find('"', NTF_href_loc2 + 8)
            NTF_href = html[NTF_href_loc2 + 6: NTF_href_loc3]
            NTF_url_link = html_link + NTF_href

            LP_href_loc1 = html.find("Legal Proceedings")
            LP_href_loc2 = html.find("href=", LP_href_loc1 - 90)
            if LP_href_loc2 > LP_href_loc1:
                LP_href_loc2 = -1
            if LP_href_loc2 == -1:
                LP_href_loc2 = html.find("HREF=", LP_href_loc1 - 90)
            LP_href_loc3 = html.find('"', LP_href_loc2 + 8)
            LP_href = html[LP_href_loc2 + 6: LP_href_loc3]
            LP_url_link = html_link + LP_href

            RF_href_loc1 = html.find("Risk Factors")
            RF_href_loc2 = html.find("href=", RF_href_loc1 - 90)
            if RF_href_loc2 > RF_href_loc1:
                RF_href_loc2 = -1
            if RF_href_loc2 == -1:
                RF_href_loc2 = html.find("HREF=", RF_href_loc1 - 90)
            RF_href_loc3 = html.find('"', RF_href_loc2 + 8)
            RF_href = html[RF_href_loc2 + 6: RF_href_loc3]
            RF_url_link = html_link + RF_href

        elif conformed_submission_type == "10-K":
            MDA_href_loc1 = html.find("Management")
            MDA_href_loc2 = html.find("href=", MDA_href_loc1 - 150)
            if MDA_href_loc2 == -1:
                MDA_href_loc2 = html.find("HREF=", MDA_href_loc1 - 150)
            MDA_href_loc3 = html.find('"', MDA_href_loc2 + 8)
            MDA_href = html[MDA_href_loc2 + 6: MDA_href_loc3]
            MDA_url_link = html_link + MDA_href

            NTF_href_loc1 = html.find("Financial Statements")
            NTF_href_loc2 = html.find("href=", NTF_href_loc1 - 150)
            if NTF_href_loc2 == -1:
                NTF_href_loc2 = html.find("HREF=", NTF_href_loc1 - 150)
            NTF_href_loc3 = html.find('"', NTF_href_loc2 + 8)
            NTF_href = html[NTF_href_loc2 + 6: NTF_href_loc3]
            NTF_url_link = html_link + NTF_href

            LP_href_loc1 = html.find("Legal Proceedings")
            LP_href_loc2 = html.find("href=", LP_href_loc1 - 150)
            if LP_href_loc2 == -1:
                LP_href_loc2 = html.find("HREF=", LP_href_loc1 - 150)
            LP_href_loc3 = html.find('"', LP_href_loc2 + 8)
            LP_href = html[LP_href_loc2 + 6: LP_href_loc3]
            LP_url_link = html_link + LP_href

            RF_href_loc1 = html.find("Risk Factors")
            RF_href_loc2 = html.find("href=", RF_href_loc1 - 150)
            if RF_href_loc2 == -1:
                RF_href_loc2 = html.find("HREF=", RF_href_loc1 - 150)
            RF_href_loc3 = html.find('"', RF_href_loc2 + 8)
            RF_href = html[RF_href_loc2 + 6: RF_href_loc3]
            RF_url_link = html_link + RF_href

            BD_href_loc1 = html.find("Business", 300)
            BD_href_loc2 = html.find("href=", BD_href_loc1 - 150)
            if BD_href_loc2 == -1:
                BD_href_loc2 = html.find("HREF=", BD_href_loc1 - 150)
            BD_href_loc3 = html.find('"', BD_href_loc2 + 8)
            BD_href = html[BD_href_loc2 + 6: BD_href_loc3]
            BD_url_link = html_link + BD_href

        ws3.cell(column=1, row=i, value=company_name)
        ws3.cell(column=2, row=i, value=conformed_submission_type)
        ws3.cell(column=3, row=i, value=filing_date)
        ws3.cell(column=4, row=i, value=conformed_period_of_report)
        ws3.cell(column=5, row=i, value=fiscal_year_end)
        ws3.cell(column=6, row=i, value=accession_number)
        ws3.cell(column=7, row=i, value=cik_number)
        ws3.cell(column=8, row=i, value=sic_group)
        ws3.cell(column=9, row=i, value=irs_number)
        ws3.cell(column=10, row=i, value=state_incorporated)
        ws3.cell(column=11, row=i, value=street)
        ws3.cell(column=12, row=i, value=city)
        ws3.cell(column=13, row=i, value=state)
        ws3.cell(column=14, row=i, value=zipcode)
        ws3.cell(column=15, row=i, value=text_url)
        ws3.cell(column=16, row=i, value=html_link)
        ws3.cell(column=17, row=i, value=MDA_url_link)
        ws3.cell(column=18, row=i, value=NTF_url_link)
        ws3.cell(column=19, row=i, value=LP_url_link)
        ws3.cell(column=20, row=i, value=RF_url_link)
        ws3.cell(column=21, row=i, value=BD_url_link)

        year = conformed_period_of_report[0:4]
        num_year = int(year)
        fin_directory = 'F:/Dropbox/IvsTbx/Financial Data/'
        date_check = ""
        if conformed_period_of_report.find("06")!=-1:
            quarter = "q3"
            date_check = year+"0630"
        elif conformed_period_of_report.find("03") != -1:
            quarter = "q2"
            date_check = year+"0331"
        elif conformed_period_of_report.find("12") != -1:
            quarter = "q1"
            date_check = year+"1231"
            num_year = num_year + 1
        elif conformed_period_of_report.find("09") != -1:
            quarter = "q4"
            date_check = year+"0930"
        year = str(num_year)
        fin_file = fin_directory+"Financial Data "+year+quarter+".txt"
        fin = open(fin_file)

        for idx,line in enumerate(fin.readlines()):
            date_info = line[line.find('us-gaap')+13:line.find('us-gaap')+22].replace(" ", "")
            if line[0:20] == accession_number and date_info.find(year)!=-1:
                ws2.cell(column=1, row=x, value=accession_number)
                tag = line[20:line.find('us-gaap')].replace(" ", "")
                tag = re.sub(r"(\w)([A-Z])", r"\1 \2", tag)
                ws2.cell(column=3, row=x, value=tag)
                ws2.cell(column=4, row=x, value=line[line.find('us-gaap'):line.find('us-gaap')+13].replace(" ", ""))
                ws2.cell(column=6, row=x, value=line[line.find('us-gaap')+13:line.find('us-gaap')+22].replace(" ", ""))
                ws2.cell(column=5, row=x, value='')
                ws2.cell(column=7, row=x, value=line[line.find('us-gaap')+22:line.find('us-gaap')+24].replace(" ", ""))
                ws2.cell(column=8, row=x, value=line[line.find('us-gaap')+24:line.find('us-gaap')+28].replace(" ", ""))
                ws2.cell(column=9, row=x, value=line[line.find('us-gaap') + 28:].replace(" ", ""))
                if ws2.cell(column = 8, row = x).value == '	sha':
                    ws2.cell(column=8, row=x, value=line[line.find('us-gaap') + 24:line.find('us-gaap') + 31].replace(" ", ""))
                    ws2.cell(column=9, row=x, value=line[line.find('us-gaap') + 31:].replace(" ", ""))
                ws2.cell(column=10, row=x, value='')
                if ws2.cell(column=3, row=x).value.find(accession_number) != -1:
                    ws2.cell(column=4, row=x, value=tag.replace(" ", ""))
                    ws2.cell(column=4, row=x, value='')
                    ws2.cell(column=5, row=x, value=accession_number)
                    ws2.cell(column=6, row=x, value=line[line.find(accession_number,20) + 20:line.find(accession_number,20)+30].replace(" ", ""))
                    ws2.cell(column=7, row=x, value=line[line.find(accession_number,20) + 31:line.find(accession_number,20)+32].replace(" ", ""))
                    if line.find('shares', line.find(accession_number,20)+28)==-1:
                        ws2.cell(column=8, row=x, value='USD')
                        ws2.cell(column=9, row=x, value=line[line.find('USD') + 4:].replace(" ", ""))
                    else:
                        ws2.cell(column=8, row=x, value='Shares')
                        ws2.cell(column=9, row=x, value=line[line.find('shares') + 7:].replace(" ", ""))
                x=x+1
        i=i+1



dest_filename = 'Financial_Output.xlsx'
wb.save(filename = dest_filename)
