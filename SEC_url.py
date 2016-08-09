import urllib2
from openpyxl import Workbook
from openpyxl import load_workbook

with open('SEC_Filing_Info.csv', 'wb') as csv_file:
    field_names = ['Company_Name', 'Submission_Type', 'Filing_Date', 'Period_of_Report', 'Fiscal_Year_End',
                   'Accession_Number', 'CIK', 'SIC', 'IRS_Number', 'State_Incorporated', 'Street', 'City', 'State',
                   'Zipcode', 'Text_URL', 'HTML_Link', 'MDA_Link', 'NTF_Link', 'LP_Link', 'RF_Link', 'BD_Link']
    #wb = Workbook()
    #ws1 = wb.active
    #ws1.title = "Filing Metadata"
    #ws1.append(field_names)
    #wb2 = load_workbook('SEC_Filing_Info.xlsx')
    for row in wb2:
        response = urllib2.urlopen(row[0])
        html = response.read()
        sec_url = 'https://www.sec.gov/Archives/edgar/data/'

        url_link = row[0]

        accession_number_loc1 = html.find("ACCESSION NUMBER:") + 19
        accession_number_loc2 = html.find("\n", accession_number_loc1)
        accession_number = html[accession_number_loc1:accession_number_loc2]

        conformed_submission_type_loc1 = html.find("CONFORMED SUBMISSION TYPE:") + 27
        conformed_submission_type_loc2 = html.find("\n", conformed_submission_type_loc1)
        conformed_submission_type = html[conformed_submission_type_loc1:conformed_submission_type_loc2]

        conformed_period_of_report_loc1 = html.find("CONFORMED PERIOD OF REPORT:") + 28
        conformed_period_of_report_loc2 = html.find("\n", conformed_period_of_report_loc1)
        conformed_period_of_report = html[conformed_period_of_report_loc1:conformed_period_of_report_loc2]

        filing_date_loc1 = html.find("FILED AS OF DATE:") + 19
        filing_date_loc2 = html.find("\n", filing_date_loc1)
        filing_date = html[filing_date_loc1:filing_date_loc2]

        company_name_loc1 = html.find("COMPANY CONFORMED NAME:") + 26
        company_name_loc2 = html.find("\n", company_name_loc1)
        company_name = html[company_name_loc1:company_name_loc2]

        cik_number_loc1 = html.find("CENTRAL INDEX KEY:") + 21
        cik_number_loc2 = html.find("\n", cik_number_loc1)
        cik_number = html[cik_number_loc1:cik_number_loc2]

        cik_url = url_link[url_link.find("data/")+5:url_link.find("/", url_link.find("data/")+5)]
        text_url = sec_url + cik_url + "/" + accession_number + ".txt"

        file_name_loc1 = html.find("<FILENAME>")+10
        file_name_loc2 = html.find("\n", file_name_loc1)
        file_name = html[file_name_loc1:file_name_loc2]

        sic_group_loc1 = html.find("STANDARD INDUSTRIAL CLASSIFICATION:") + 36
        sic_group_loc2 = html.find("\n", sic_group_loc1)
        sic_group = html[sic_group_loc1:sic_group_loc2]

        irs_number_loc1 = html.find("IRS NUMBER:") + 15
        irs_number_loc2 = html.find("\n", irs_number_loc1)
        irs_number = html[irs_number_loc1:irs_number_loc2]

        state_incorporated_loc1 = html.find("STATE OF INCORPORATION:") + 26
        state_incorporated_loc2 = html.find("\n", state_incorporated_loc1)
        state_incorporated = html[state_incorporated_loc1:state_incorporated_loc2]

        fiscal_year_end_loc1 = html.find("FISCAL YEAR END:") + 19
        fiscal_year_end_loc2 = html.find("\n", fiscal_year_end_loc1)
        fiscal_year_end = html[fiscal_year_end_loc1:fiscal_year_end_loc2]

        street_loc1 = html.find("STREET 1:") + 11
        street_loc2 = html.find("\n", street_loc1)
        street = html[street_loc1:street_loc2]

        city_loc1 = html.find("CITY:") + 8
        city_loc2 = html.find("\n", city_loc1)
        city = html[city_loc1:city_loc2]

        state_loc1 = html.find("STATE:") + 9
        state_loc2 = html.find("\n", state_loc1)
        state = html[state_loc1:state_loc2]

        zip_loc1 = html.find("ZIP:") + 7
        zip_loc2 = html.find("\n", zip_loc1)
        zipcode = html[zip_loc1:zip_loc2]

        sec_txt = open(company_name.replace('/', '')+' '+conformed_submission_type+' '+filing_date+'.txt', 'w+')
        sec_txt.write(html)
        sec_txt.close()

        html_link = sec_url + cik_url + "/" + accession_number.replace("-", "") + "/" + file_name

        MDA_url_link = ""
        NTF_url_link = ""
        RF_url_link = ""
        LP_url_link = ""
        BD_url_link = ""

        print company_name
        print conformed_submission_type

        if conformed_submission_type == "10-Q":
            MDA_href_loc1 = html.find("Management&")
            MDA_href_loc2 = html.find("href=", MDA_href_loc1-90)
            if MDA_href_loc2 > MDA_href_loc1:
                MDA_href_loc2 = -1
            if MDA_href_loc2 == -1:
                MDA_href_loc2 = html.find("HREF=", MDA_href_loc1-90)
            MDA_href_loc3 = html.find('"', MDA_href_loc2 + 8)
            MDA_href = html[MDA_href_loc2 + 6: MDA_href_loc3]
            MDA_url_link = html_link+MDA_href

            NTF_href_loc1 = html.find("Notes to Con")
            NTF_href_loc2 = html.find("href=", NTF_href_loc1-90)
            if NTF_href_loc2 > NTF_href_loc1:
                NTF_href_loc2 = -1
            if NTF_href_loc2 == -1:
                NTF_href_loc2 = html.find("HREF=", NTF_href_loc1 - 90)
            NTF_href_loc3 = html.find('"', NTF_href_loc2 + 8)
            NTF_href = html[NTF_href_loc2 + 6: NTF_href_loc3]
            NTF_url_link = html_link+NTF_href

            LP_href_loc1 = html.find("Legal Proceedings")
            LP_href_loc2 = html.find("href=", LP_href_loc1-90)
            if LP_href_loc2 > LP_href_loc1:
                LP_href_loc2 = -1
            if LP_href_loc2 == -1:
                LP_href_loc2 = html.find("HREF=", LP_href_loc1 - 90)
            LP_href_loc3 = html.find('"', LP_href_loc2 + 8)
            LP_href = html[LP_href_loc2 + 6: LP_href_loc3]
            LP_url_link = html_link+LP_href

            RF_href_loc1 = html.find("Risk Factors")
            RF_href_loc2 = html.find("href=", RF_href_loc1-90)
            if RF_href_loc2 > RF_href_loc1:
                RF_href_loc2 = -1
            if RF_href_loc2 == -1:
                RF_href_loc2 = html.find("HREF=", RF_href_loc1 - 90)
            RF_href_loc3 = html.find('"', RF_href_loc2 + 8)
            RF_href = html[RF_href_loc2 + 6: RF_href_loc3]
            RF_url_link = html_link+RF_href

        elif conformed_submission_type == "10-K":
            MDA_href_loc1 = html.find("Management")
            MDA_href_loc2 = html.find("href=", MDA_href_loc1 - 150)
            if MDA_href_loc2 == -1:
                MDA_href_loc2 = html.find("HREF=", MDA_href_loc1 - 150)
            MDA_href_loc3 = html.find('"', MDA_href_loc2+8)
            MDA_href = html[MDA_href_loc2 + 6: MDA_href_loc3]
            MDA_url_link = html_link + MDA_href

            NTF_href_loc1 = html.find("Financial Statements")
            NTF_href_loc2 = html.find("href=", NTF_href_loc1 - 150)
            if NTF_href_loc2 == -1:
                NTF_href_loc2 = html.find("HREF=", NTF_href_loc1 - 150)
            NTF_href_loc3 = html.find('"', NTF_href_loc2+8)
            NTF_href = html[NTF_href_loc2 + 6: NTF_href_loc3]
            NTF_url_link = html_link + NTF_href

            LP_href_loc1 = html.find("Legal Proceedings")
            LP_href_loc2 = html.find("href=", LP_href_loc1 - 150)
            if LP_href_loc2 == -1:
                LP_href_loc2 = html.find("HREF=", LP_href_loc1 - 150)
            LP_href_loc3 = html.find('"', LP_href_loc2+8)
            LP_href = html[LP_href_loc2 + 6: LP_href_loc3]
            LP_url_link = html_link + LP_href

            RF_href_loc1 = html.find("Risk Factors")
            RF_href_loc2 = html.find("href=", RF_href_loc1 - 150)
            if RF_href_loc2 == -1:
                RF_href_loc2 = html.find("HREF=", RF_href_loc1 - 150)
            RF_href_loc3 = html.find('"', RF_href_loc2+8)
            RF_href = html[RF_href_loc2 + 6: RF_href_loc3]
            RF_url_link = html_link + RF_href

            BD_href_loc1 = html.find("Business", 300)
            BD_href_loc2 = html.find("href=", BD_href_loc1 - 150)
            if BD_href_loc2 == -1:
                BD_href_loc2 = html.find("HREF=", BD_href_loc1 - 150)
            BD_href_loc3 = html.find('"', BD_href_loc2 + 8)
            BD_href = html[BD_href_loc2 + 6: BD_href_loc3]
            BD_url_link = html_link + BD_href

        writer.writerow({'Company_Name': company_name, 'Submission_Type': conformed_submission_type, 'Filing_Date': filing_date, 'Period_of_Report': conformed_period_of_report, 'Fiscal_Year_End': fiscal_year_end,
                         'Accession_Number': accession_number, 'CIK': cik_number, 'SIC': sic_group, 'IRS_Number': irs_number, 'State_Incorporated': state_incorporated, 'Street': street, 'City': city, 'State': state,
                         'Zipcode': zipcode, 'Text_URL': text_url, 'HTML_Link': html_link, 'MDA_Link': MDA_url_link, 'NTF_Link': NTF_url_link, 'LP_Link': LP_url_link, 'RF_Link': RF_url_link, 'BD_Link': BD_url_link})

    csv_file.close()

